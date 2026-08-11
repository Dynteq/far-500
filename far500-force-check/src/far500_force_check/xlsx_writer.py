"""Bouw het XLSX-rapport (setup_analyse + data) uit een Analysis."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from . import charts
from .model import Analysis

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITLE_FONT = Font(size=16, bold=True)
HEADER_FONT = Font(bold=True)
VERDICT_FONT_PASS = Font(size=28, bold=True, color="006100")
VERDICT_FONT_FAIL = Font(size=28, bold=True, color="9C0006")

DATA_COLUMNS = [
    "t_s", "angle_deg", "force_N", "arc_cm", "speed_cm_s", "accel_cm_s2", "handle_height_cm",
    "region", "limit_N", "in_grace", "force_ok",
]


def build_workbook(analysis: Analysis, include_data_tab: bool = True) -> Workbook:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "setup_analyse"
    _write_setup_analyse(ws1, analysis)
    if include_data_tab:
        ws2 = wb.create_sheet("data")
        _write_data_sheet(ws2, analysis)
    return wb


def _write_setup_analyse(ws: Worksheet, analysis: Analysis) -> None:
    meta = analysis.recording.meta
    criteria = analysis.criteria
    row = 1

    ws.cell(row=row, column=1, value="FAR-500 — bedienkracht-analyse").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=meta.naam or "(naam onbekend)")
    row += 1
    if meta.notities:
        ws.cell(row=row, column=1, value=meta.notities)
        row += 1
    ws.cell(row=row, column=1, value=meta.datum_tijd or "")
    row += 2

    ws.cell(row=row, column=1, value="Setup").font = HEADER_FONT
    row += 1
    setup_rows = [
        ("Meetas", meta.raw_extra.get("Meetas", meta.kalibratie)),
        ("L ingevoerd / berekend / gebruikt (cm)", f"{meta.l_ingevoerd_cm} / {meta.l_berekend_cm} / {meta.l_used}"),
        ("H ingevoerd / berekend / gebruikt (cm)", f"{meta.h_ingevoerd_cm} / {meta.h_berekend_cm} / {meta.h_used}"),
        ("Handvat hoog (cm) / hoek (deg)", f"{meta.handvat_hoog_cm} / {meta.hoek_hoog_deg}"),
        ("Handvat laag (cm) / hoek (deg)", f"{meta.handvat_laag_cm} / {meta.hoek_laag_deg}"),
        ("Max snelheid (cm/s)", meta.max_snelheid_cm_s if meta.max_snelheid_cm_s is not None else "(default 20)"),
        ("Max versnelling (cm/s2)", meta.max_versnelling_cm_s2 if meta.max_versnelling_cm_s2 is not None else "(default 40)"),
    ]
    for label, value in setup_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=str(value))
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Criteria").font = HEADER_FONT
    row += 1
    criteria_text = [
        f"C1 — onder {criteria.H_THRESH:.0f} cm handvathoogte: kracht ≤ {criteria.F_LOW:.0f} N "
        f"(eerste {criteria.GRACE_ARC:.0f} cm vanaf het anker: ≤ {criteria.GRACE_FACT * criteria.F_LOW:.1f} N breakaway-marge).",
        f"C2 — boven {criteria.H_THRESH:.0f} cm handvathoogte: kracht ≤ {criteria.F_HIGH:.0f} N "
        f"(zelfde breakaway-marge tot {criteria.GRACE_FACT * criteria.F_HIGH:.1f} N, A1 GRACE_ABOVE={criteria.GRACE_ABOVE}).",
        f"C3 — handvathoogte ≤ {criteria.H_MAX:.0f} cm (A3 HEIGHT_MAX_IS_FAIL={criteria.HEIGHT_MAX_IS_FAIL}).",
        "C4 — snelheid ≤ Max snelheid, versnelling ≤ Max versnelling (uit de metadata, anders default 20/40). "
        f"Informatief: geen officiële eis, telt {'wél' if criteria.C4_AFFECTS_OVERALL else 'niet'} mee in het "
        "eindoordeel (A6 C4_AFFECTS_OVERALL).",
        f"Anker: overschrijdt |F| de basislimiet binnen de eerste {criteria.DETECT_ARC:.0f} cm van een beweging, "
        "dan ligt het anker op dat kruisingspunt; anders op het bewegingsbegin. Het "
        f"{criteria.GRACE_ARC:.0f} cm-breakaway-venster geldt in beide gevallen vanaf dat anker (zie README).",
        f"Kracht getoetst als |force_N| (A2, USE_ABS_FORCE={criteria.USE_ABS_FORCE}). "
        f"Richtingsdetectie met REV_HYST={criteria.REV_HYST:.1f} cm hysterese.",
    ]
    for line in criteria_text:
        cell = ws.cell(row=row, column=1, value=line)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Uitkomst").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Criterium").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Resultaat").font = HEADER_FONT
    ws.cell(row=row, column=3, value="Onderbouwing").font = HEADER_FONT
    row += 1
    outcome_rows = [
        ("C1", analysis.c1_ok, f"hoogste kracht onder {criteria.H_THRESH:.0f} cm = {analysis.peak_force_low_N:.1f} N ≤ {criteria.F_LOW:.0f}/{criteria.GRACE_FACT*criteria.F_LOW:.1f} N"),
        ("C2", analysis.c2_ok, f"hoogste kracht boven {criteria.H_THRESH:.0f} cm = {analysis.peak_force_high_N:.1f} N ≤ {criteria.F_HIGH:.0f}/{criteria.GRACE_FACT*criteria.F_HIGH:.1f} N"),
        ("C3", analysis.c3_ok, f"max hoogte = {analysis.max_height_cm:.1f} cm (limiet {criteria.H_MAX:.0f} cm)"),
        (
            "C4 (info)",
            analysis.c4_ok,
            f"piek snelheid = {analysis.max_speed_cm_s:.1f} cm/s, piek versnelling = {analysis.max_accel_cm_s2:.1f} cm/s2"
            + ("" if criteria.C4_AFFECTS_OVERALL else " — geen officiële eis, telt niet mee in het eindoordeel"),
        ),
    ]
    for label, ok, note in outcome_rows:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value="PASS" if ok else "FAIL")
        cell.fill = GREEN_FILL if ok else RED_FILL
        cell.font = HEADER_FONT
        ws.cell(row=row, column=3, value=note)
        row += 1
    row += 1

    verdict_cell = ws.cell(row=row, column=1, value="EINDOORDEEL: " + ("PASS" if analysis.overall_pass else "FAIL"))
    verdict_cell.font = VERDICT_FONT_PASS if analysis.overall_pass else VERDICT_FONT_FAIL
    verdict_cell.fill = GREEN_FILL if analysis.overall_pass else RED_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 36
    row += 2

    chart_anchor_row = row
    helper_start_row = row + 22
    next_helper_row = charts.build_force_chart(
        ws, analysis, start_row=helper_start_row, start_col=10, anchor_cell=f"A{chart_anchor_row}"
    )
    if any(s.handle_height_cm is not None for s in analysis.recording.samples):
        charts.build_height_chart(
            ws, analysis, start_row=next_helper_row, start_col=10, anchor_cell=f"A{chart_anchor_row + 19}"
        )
        row = chart_anchor_row + 38
    else:
        row = chart_anchor_row + 19

    for col, width in zip("ABCDEF", (36, 16, 60, 16, 16, 16)):
        ws.column_dimensions[col].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:F{row}"


def _write_data_sheet(ws: Worksheet, analysis: Analysis) -> None:
    for j, h in enumerate(DATA_COLUMNS, start=1):
        ws.cell(row=1, column=j, value=h).font = HEADER_FONT

    for i, (s, r) in enumerate(zip(analysis.recording.samples, analysis.sample_results), start=2):
        ws.cell(row=i, column=1, value=s.t_s)
        ws.cell(row=i, column=2, value=s.angle_deg)
        ws.cell(row=i, column=3, value=s.force_N)
        ws.cell(row=i, column=4, value=s.arc_cm)
        ws.cell(row=i, column=5, value=s.speed_cm_s)
        ws.cell(row=i, column=6, value=s.accel_cm_s2)
        ws.cell(row=i, column=7, value=s.handle_height_cm)
        ws.cell(row=i, column=8, value=r.region)
        ws.cell(row=i, column=9, value=r.envelope_N)
        ws.cell(row=i, column=10, value=r.in_grace)
        ws.cell(row=i, column=11, value=r.force_ok)

    ws.freeze_panes = "A2"
    last_row = 1 + len(analysis.recording.samples)
    ws.auto_filter.ref = f"A1:{chr(ord('A') + len(DATA_COLUMNS) - 1)}{last_row}"
